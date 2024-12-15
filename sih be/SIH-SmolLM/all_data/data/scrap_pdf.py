import json
import os
import pdfplumber
import pandas as pd
from openpyxl import Workbook

def excel(pdf_files, years, location):
    output_path = os.path.join(location,f"{years}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "College Data"

    ws.cell(row=1, column=1, value="College")
    current_row = 2 
    for college_data in pdf_files:
        pdf_path = college_data["file"]
        college_name = college_data["college"]
        ws.cell(row=current_row, column=1, value=college_name)
        current_col = 2

        with pdfplumber.open(pdf_path) as pdf:
            max_row_used = current_row     
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                if tables:
                    for table_index, table in enumerate(tables):
                        try:
                            df = pd.DataFrame(table[1:], columns=table[0])
                        except Exception:
                            continue 
                        
                        start_row = current_row
                        start_col = current_col
                        
                        for col_index, header in enumerate(df.columns, start=start_col):
                            ws.cell(row=start_row, column=col_index, value=header)
                        
                        for row_index, row in enumerate(df.itertuples(index=False), start=start_row + 1):
                            for col_index, value in enumerate(row, start=start_col):
                                ws.cell(row=row_index, column=col_index, value=value)
                        
                        max_row_used = max(max_row_used, start_row + len(df))
                        current_col += len(df.columns) + 1
            current_row = max_row_used + 2
    wb.save(output_path)
    print(f"Model saved for {year}")

root_dir = "data"

years = [str(year) for year in range(2017, 2025)]

pdf_list = []

for main_category in os.listdir(root_dir):
    main_category_path = os.path.join(root_dir, main_category)
    print(main_category_path)
    if os.path.isdir(main_category_path): 
        for year in years:
            year_dict = {"year": year, "files": []}  
            for college in os.listdir(main_category_path):  
                college_path = os.path.join(main_category_path, college)
                if os.path.isdir(college_path):  
                    pdf_path = os.path.join(college_path, f"{year}.pdf")
                    if os.path.isfile(pdf_path):  
                        year_dict["files"].append({
                            "file": pdf_path,
                            "college": college,
                            "category": main_category  
                        })
            pdf_files = year_dict['files']
            excel(pdf_files, year, main_category_path)



output_file = "pdf_list.json"
with open(output_file, "w") as f:
    json.dump(pdf_list, f, indent=4)